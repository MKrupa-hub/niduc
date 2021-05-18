from openpyxl import Workbook
import xmlHandler as hand
import Generator as gene
import Decoder as deco
import csv
import random

def write_result(dest_fileName, type):
    result = []
    for row in range(sample):
        temp = [0, 0, 0, 0]
        for j in range(test):
            packet = gene.packing(length)
            if type == 0:
                coded = gene.code_hamming(packet)
            if type == 1:
                coded = gene.multiple_bit(multi, packet)
            if type == 2:
                coded = gene.code_crc(packet)
            for i in range(len(coded)):
                if random.random() < probability:
                    coded[i] = int(not coded[i])
            if type == 0:
                decoded, fixed = deco.decode_hamming(coded)
            if type == 1:
                decoded, fixed = deco.decodeMulti(coded, multi)
            if type == 2:
                decoded, fixed = deco.decodeCrc(coded)
            # Musimy teraz ocenic ile pakietow zostalo przeslanych bez bledow, ile mialo bledy i czy udalo sie je naprawic. ORAZ ile niewykrytych bledow wystapilo
            if fixed:
                # Wystapila naprawa przy dekoderze. Sprawdzam czy naprawil poprawnie. Robie to przez porownanie listy przed i po wysyle
                if packet == decoded:
                    temp[1] = temp[1] + 1
                else:
                    temp[2] = temp[2] + 1
            else:
                # Nie wystapila naprawa ale moze wystapil niewykryty blada
                if packet == decoded:
                    temp[0] = temp[0] + 1
                else:
                    temp[3] = temp[3] + 1
        result.append(temp)
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheetNames[0]
    sheet.cell(row=1, column=1).value = columnname[0]
    sheet.cell(row=1, column=2).value = columnname[1]
    sheet.cell(row=1, column=3).value = columnname[2]
    sheet.cell(row=1, column=4).value = columnname[3]
    for row in range(2, sample + 2):
        for column in range(1, 5):
            sheet.cell(column = column, row = row).value = result[row - 2][column - 1]
    wb.create_sheet(sheetNames[1], 1)
    hand.calculateAverage(wb, sheetNames, columnname, sample)
    wb.save(dest_fileName)

sheetNames = [ 'Sim', 'Calc']
columnname = ['Good', 'Fixed', 'Fixed not', 'Undetected']
fileNames = ['_hamming.xlsx', '_multi.xlsx', '_crc.xlsx']
multi = 4 # ile bitow powtarzany
length = 11 # dlugosc pakietu
sample = 100 # ile wysylow wykonuje
test = 1000 # Ile pakietow wysylamy
probability = 0.05 # prawdopodobienstwo przeklamania bitu

# sample, test, fileNames, sheetNames
# handler = Handler(sample, test, fileNames, sheetNames)
# handler.calculateAverage()
# Kanal transmisyjny BSC
# mamy zakodowany pakiet, przysylamy go przez kanal
# jest prawdopodobienstwo ze zamienimy bit w pakiecie na przeciwny
write_result(fileNames[0], 0)
# write_result(fileNames[1], 1)
# write_result(fileNames[2], 2)

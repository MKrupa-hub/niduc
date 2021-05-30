import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook
import math
from pathlib import Path
import statistics
from scipy.stats import norm
import scipy.optimize as opt
import matplotlib.mlab as mlab
import warnings
import os
from openpyxl.drawing.image import Image

message = ['Min value', '25th percentile', '50th percentile', '75th percentile', 'Max value']
imgWidth = 390
imgHeight = 450
charts = "charts"
excelSam = "simResult"

def pdf(x,a, mu, sigma):
    return a* math.e ** ((-1 / 2) * (((x - mu) / sigma) ** 2))

def printColumn(start, sheet, columnname):
    for i in range(start,start+4):
        sheet.cell(row=1, column=i).value = columnname[i - start]

def handle(sheetNames, columnname, sample, result, dest_fileName):
    warnings.filterwarnings('ignore')
    cwd = os.getcwd()
    Path(cwd + '/' + charts).mkdir(parents = True, exist_ok = True)
    Path(cwd + '/' + excelSam).mkdir(parents=True, exist_ok = True)
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheetNames[0]
    printColumn(1, sheet, columnname)
    printColumn(7, sheet, columnname)
    printColumn(14, sheet, columnname)
    for row in range(2, sample + 2):
        for column in range(1, 5):
            sheet.cell(column=column, row=row).value = result[row - 2][column - 1]

    sheet['F2'] = 'Average'
    sheet['F3'] = 'Standard dev.'
    sheet['F5'] = 'Quartile'
    sheet['M2'] = 'a'
    sheet['M3'] = 'mu'
    sheet['M4'] = 'sigma'
    for i in range(5):
        sheet['F' + str((6 + i))] = i
        sheet['K' + str((6 + i))] = message[i]

    for iter in range(len(columnname)):
        data = []
        [data.append(result[x][iter]) for x in range(sample)]
        mean = statistics.mean(data)
        std = statistics.stdev(data)
        quartiles = statistics.quantiles(data, n=6, method='inclusive')
        iqr = quartiles[3] - quartiles[1]
        box = plt.boxplot(quartiles, vert=False)
        plt.title(columnname[iter])
        plt.xlabel('Liczba pakietów przesłanych')
        plt.savefig(charts + '/' + dest_fileName[1] + ' ' + columnname[iter] + 'box.jpg')
        plt.close()
        img = Image(charts + '/' + dest_fileName[1] + ' ' + columnname[iter] + 'box.jpg')
        img.height = imgHeight
        img.width = imgWidth
        sheet.add_image(img, ('H' + str(((iter+1) * 20))))

        counts, bins, bars = plt.hist(data, bins=20)
        x = []
        for i in range(len(bins) - 1):
            x.append((bins[i] + bins[i + 1]) / 2)
        y = counts
        try:
            params, pcov = opt.curve_fit(pdf, x, y, p0=[max(y), quartiles[2], iqr / 1.349], maxfev=5000)
            plt.plot(x, pdf(x, params[0], params[1], params[2]))
            plt.title(columnname[iter])
            for i in range(len(params)):
                sheet[chr(78+iter) + str((2+i))] = params[i]
        except RuntimeError:
            plt.title(columnname[iter] + "couldn't estimate Gauss")
        finally:
            plt.xlabel('Liczba pakietów przesłanych')
            plt.ylabel('Liczba wystąpień')
            plt.savefig(charts + '/' + dest_fileName[1] + ' ' + columnname[iter] + 'hist.jpg')
            plt.close() 
            img = Image(charts + '/' + dest_fileName[1] + ' ' + columnname[iter] + 'hist.jpg')
            img.height = imgHeight
            img.width = imgWidth
            sheet.add_image(img, ('P' + str(((iter +1) * 20))))

            sheet[chr(71 + iter) + '2'] = mean
            sheet[chr(71 + iter) + '3'] = std

            for i in range(len(quartiles)):
                sheet[chr(71 + iter) + str((6+i))] = quartiles[i]

    wb.save(excelSam + '/' + dest_fileName)
